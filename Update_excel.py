import requests
import urllib
from urllib.request import urlopen
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook

Server = ("스카니아", "루나", "크로아", "베라","엘리시움","오로라","리부트","이노시스")

Member = (("권강현","밍먕의",0),("김대성","머성머",0),("김민섭","모법또비",6)
        ,("★김민식★","SiGy",3),("김민주","헬로후닝",1),("김상완","얼렁",1),("김성원","fims13",6),
        ("김성은","민들레u",6),("김재홍","JH메신",3),("김정현","뚠두붕",1),("김준수","죽염과송염",1),
        ("김준엽","엽쭌이",4),("김지호","별빛사서",1),("류한일","죤방",6),("문기봉","춤추는이병",1),
        ("박선호","슈리언즈",0),("박정혁","그러죠뭐",5),("박정환","하늘다람쥐다",1),("박지호","SHERO",1),
        ("박진제","뿌쟝",4),("변형각","맹구포드",1),("빈석현","너와함께랄라",6),
        ("서동진","땡진7번",0),("서효빈","음악한모금",4),("손민재","신징",2),("손정수","얍힝",0),
        ("신제철","박사할걸",1),("안성현","배틀메이지",2),("양수정","권투초보",3),("원동연","도치츄",2),("윤주희","계란탕라면",1),
        ("이누리","아힌느",7),("이재복","꽃촬",2),("이철희","YYeah",0),
        ("이현식","이현식",1),("임승혁","존예여신연희",0),("장요셉","장요셉",1),("전일우","쿠키선비",0),
        ("전채연","석사낙지",1),("정종한","온pv",3),("조인수","토스트나로오",1),
        ("최민호","쪼꼬",1),("최재원","하늘뱀장어",1),("최준용","람검",4),("한민욱","쓸윈부쓸윈브",1),
        ("한승수","티모선",0))

load_wb=load_workbook("Weekly_Report.xlsx",data_only=True) 
sheet= load_wb['Sheet1'] #기존 엑셀파일 오픈하여 지난주 유니온 / 레벨 수치 추출

prev_level=[]
prev_union=[]

for i in range(len(Member)):
    prev_level.append(int(sheet.cell(row=i+2,column=5).value))
    try:
        prev_union.append(int(sheet.cell(row=i+2,column=6).value))
    except:
        prev_union.append('-')




Rnum=2 #첫번째 Row는 Col명, 2번째 Row부터 시작

for Name, Nickname, Snum in Member:
    Convert_Nick=urllib.parse.quote(Nickname)
    Page = urlopen(f"https://maple.gg/u/{Convert_Nick}")
    soup = bs(Page,"html.parser")

    #무릉 정보
    try:
        tower = soup.body.find("h1",{"class":"user-summary-floor font-weight-bold"}).text
        Tower = int(tower.split("\n")[0])
    except:
        Tower ="-"
    #유니온
    try:
        union = soup.body.find("span",{"class":"user-summary-level"}).text
        Union=int(union.split(".")[1])
    except:
        Union ="-"

    #캐릭터 정보 (레벨, 직업)
    character = soup.body.find("ul",{"class":"user-summary-list"}).text
    Character = character.split("\n")
    
    Level = int(Character[1][3]+Character[1][4]+Character[1][5])

    Occupation = Character[2]

    #마지막 접속일
    try:
        lastupdate=soup.body.find("span",{"class":"font-size-12 text-white"}).text
        lastUpdate = lastupdate.replace(" ","")
        Lastupdate = lastUpdate.replace("\n","")
        LastUpdate = Lastupdate[7]+Lastupdate[8]+Lastupdate[9]
    except:
        LastUpdate = "-"

    #print(Server[Snum],Nickname,Occupation,Level,Tower,Union)
    sheet.cell(row=Rnum,column=1).value=Name
    sheet.cell(row=Rnum,column=2).value=Nickname
    sheet.cell(row=Rnum,column=3).value=Server[Snum]
    sheet.cell(row=Rnum,column=4).value=Occupation
    sheet.cell(row=Rnum,column=5).value=Level
    sheet.cell(row=Rnum,column=6).value=Union
    sheet.cell(row=Rnum,column=7).value=Tower

    try:
        sheet.cell(row=Rnum,column=8).value=Union-prev_union[Rnum-2]
    except:
        sheet.cell(row=Rnum,column=8).value='-'

    sheet.cell(row=Rnum,column=9).value=LastUpdate
    
    tmp = Level - prev_level[Rnum-2] 

    if tmp > 0:
        sheet.cell(row=Rnum,column=10).value='O'
    else:
        sheet.cell(row=Rnum,column=10).value='X'

    Rnum=Rnum+1

load_wb.save("Weekly_Report.xlsx")
